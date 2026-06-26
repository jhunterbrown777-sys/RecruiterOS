from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK_PATH = Path("data/JOB_TRACKER_AGENT_MVP.xlsx")


def load_tracker():
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb.worksheets[0]
    return wb, ws


def get_header_map(ws):
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def find_next_pending_job(ws):
    headers = get_header_map(ws)

    job_desc_col = headers["Job Description"]
    fit_score_col = headers["Fit Score"]

    for row in range(2, ws.max_row + 1):
        job_description = ws.cell(row=row, column=job_desc_col).value
        fit_score = ws.cell(row=row, column=fit_score_col).value

        if job_description and not fit_score:
            return row

    return None


def read_job(ws, row):
    headers = get_header_map(ws)

    fields = [
        "Job ID",
        "Priority",
        "Job Title",
        "Company",
        "Location",
        "Work Arrangement",
        "Job URL",
        "Job Description",
        "Source",
    ]

    return {
        field: ws.cell(row=row, column=headers[field]).value
        for field in fields
        if field in headers
    }


def update_job_analysis(ws, row, analysis):
    headers = get_header_map(ws)

    updates = {
        "Fit Score": analysis.fit_score,
        "Apply Decision": analysis.apply_decision,
        "Reason": analysis.reason,
        "Missing Skills": analysis.missing_skills,
        "Tailored Resume Notes": analysis.tailored_resume_notes,
        "Cover Letter Draft": analysis.cover_letter_draft,
        "Recruiter Message": analysis.recruiter_message,
        "Next Action": analysis.next_action,
        "Last Updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    for column_name, value in updates.items():
        if column_name in headers:
            ws.cell(row=row, column=headers[column_name]).value = value


def save_tracker(wb):
    wb.save(WORKBOOK_PATH)