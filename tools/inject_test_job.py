from pathlib import Path
from openpyxl import load_workbook

WORKBOOK_PATH = Path("data/JOB_TRACKER_AGENT_MVP.xlsx")

TEST_JOB_DESCRIPTION = """
Job Title: Service Desk Analyst

Location: Remote

We are hiring a Service Desk Analyst to provide Tier 1 technical support to internal users. This role will troubleshoot Windows laptops, Microsoft 365, VPN connectivity, password resets, printers, and basic networking issues.

Responsibilities:
- Respond to tickets through a help desk system
- Troubleshoot Windows, Microsoft 365, VPN, and hardware issues
- Assist with onboarding and offboarding users
- Document solutions in the knowledge base
- Escalate complex issues to Tier 2 support
- Provide excellent customer service

Requirements:
- Strong communication and customer service skills
- Basic Windows troubleshooting
- Familiarity with TCP/IP, DNS, DHCP, and VPN concepts
- Basic Active Directory knowledge
- Strong documentation skills

Preferred:
- Security+, Network+, or A+
- Experience with ServiceNow or Jira
- Microsoft 365 administration exposure

Salary: $60,000 - $75,000
"""


def get_header_map(ws):
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def main():
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb.worksheets[0]
    headers = get_header_map(ws)

    row = ws.max_row + 1

    values = {
        "Job ID": f"TEST-{row}",
        "Priority": "High",
        "Job Title": "Service Desk Analyst",
        "Company": "Test Company",
        "Location": "Remote",
        "Work Arrangement": "Remote",
        "Job URL": "https://example.com/test-service-desk-analyst",
        "Job Description": TEST_JOB_DESCRIPTION,
        "Source": "Manual Test"
    }

    for column_name, value in values.items():
        if column_name in headers:
            ws.cell(row=row, column=headers[column_name]).value = value

    wb.save(WORKBOOK_PATH)
    print(f"Injected test job into row {row}")


if __name__ == "__main__":
    main()